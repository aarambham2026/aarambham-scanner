import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def generate_sample_excel(filename="sample_students.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aarambham Registrations"

    # Headers
    headers = ["Roll No", "Name", "Registered"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Specified roll number ranges:
    # 1. NC.AI.U4AID24001 to NC.AI.U4AID24067 (67 students)
    # 2. NC.SC.U4CSE24001 to NC.SC.U4CSE24055 (55 students)
    # 3. NC.SC.U4CSE24101 to NC.SC.U4CSE24156 (56 students)
    # 4. NC.SC.U4CSE24201 to NC.SC.U4CSE24257 (57 students)
    # 5. NC.EN.U4ECE24001 to NC.EN.U4ECE24050 (50 students)
    branches = [
        ("NC.AI.U4AID24", 1, 67, "AID"),
        ("NC.SC.U4CSE24", 1, 55, "CSE Sec A"),
        ("NC.SC.U4CSE24", 101, 156, "CSE Sec B"),
        ("NC.SC.U4CSE24", 201, 257, "CSE Sec C"),
        ("NC.EN.U4ECE24", 1, 50, "ECE")
    ]

    rows = []
    global_idx = 1
    for prefix, start_num, end_num, desc in branches:
        for num in range(start_num, end_num + 1):
            roll = f"{prefix}{num:03d}"
            name = f"Student {desc} {num:03d}"
            reg = "YES"
            rows.append((roll, name, reg))
            global_idx += 1

    for r in rows:
        ws.append(r)

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15

    wb.save(filename)
    print(f"Successfully generated {filename} with {len(rows)} Aarambham student registration records.")

if __name__ == "__main__":
    generate_sample_excel()
