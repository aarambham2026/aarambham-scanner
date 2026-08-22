import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session
from app.models import Student

def parse_and_import_excel(db: Session, file_bytes: bytes, replace_all: bool = False) -> dict:
    """
    Parses uploaded Excel file and upserts/replaces students in registrations table.
    Supports 2-column format: Roll No | Paid (Y/N) or standard formats with headers.
    """
    if replace_all:
        db.query(Student).delete()
        db.commit()

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {"success": False, "message": "Uploaded Excel file is empty."}

    # Inspect first row to detect headers
    first_row = [str(cell or "").strip() for cell in rows[0]]
    first_row_lower = [c.lower() for c in first_row]

    has_header = False
    roll_col = 0
    paid_col = 1
    name_col = None

    for idx, h in enumerate(first_row_lower):
        if h in ["roll no", "roll number", "rollno", "roll_number", "roll", "roll_no"]:
            roll_col = idx
            has_header = True
        elif h in ["paid", "registered", "registration", "is_registered", "status", "opted", "lunch opted"]:
            paid_col = idx
            has_header = True
        elif h in ["name", "student name", "student_name"]:
            name_col = idx
            has_header = True

    data_rows = rows[1:] if has_header else rows

    added = 0
    updated = 0
    errors = 0
    seen_rolls = set()

    for row in data_rows:
        if not row or not any(row):
            continue

        try:
            roll = str(row[roll_col] if roll_col < len(row) else "").strip()
            if not roll:
                errors += 1
                continue

            # Prevent duplicate processing within same file
            if roll.upper() in seen_rolls:
                continue
            seen_rolls.add(roll.upper())

            name = str(row[name_col]).strip() if (name_col is not None and name_col < len(row) and row[name_col]) else roll

            paid_val = ""
            if paid_col < len(row) and row[paid_col] is not None:
                paid_val = str(row[paid_col]).strip().upper()

            is_paid = paid_val in ["Y", "YES", "TRUE", "1", "REGISTERED", "PAID"]

            existing = db.query(Student).filter(Student.roll_number == roll).first()

            if existing:
                existing.name = name
                existing.token = roll
                existing.registered = is_paid
                updated += 1
            else:
                student = Student(
                    roll_number=roll,
                    name=name,
                    token=roll,
                    registered=is_paid,
                    entry_time=None,
                    exit_time=None
                )
                db.add(student)
                added += 1

        except Exception:
            errors += 1

    db.commit()
    return {
        "success": True,
        "added": added,
        "updated": updated,
        "errors": errors,
        "message": f"Import complete! Added: {added}, Updated: {updated}, Errors/Skipped: {errors}"
    }


def generate_excel_export(db: Session) -> io.BytesIO:
    """
    Exports all student registrations and attendance data to an Excel workbook (.xlsx).
    """
    students = db.query(Student).order_by(Student.roll_number.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aarambham Attendance Report"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")

    headers = ["Roll No", "Name", "Registration Status", "Entry Time", "Exit Time", "Status"]
    ws.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    from app.models import parse_and_format_ist

    for s in students:
        entry_str = parse_and_format_ist(s.entry_time) or "—"
        exit_str = parse_and_format_ist(s.exit_time) or "—"
        row = [
            s.roll_number,
            s.name,
            "REGISTERED" if s.registered else "NOT REGISTERED",
            entry_str,
            exit_str,
            s.status
        ]
        ws.append(row)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
